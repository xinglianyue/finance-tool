
import openpyxl
import json

file_path = '外卖账单20260527.xlsx'

wb = openpyxl.load_workbook(file_path)
print(f"工作表名称: {wb.sheetnames}")

ws = wb.active
print(f"工作表标题: {ws.title}")
print("\n前15行数据:")

data = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i &lt;= 15:
        print(f"第{i}行: {row}")
    data.append(row)

print(f"\n总行数: {len(data)}")

# 检查第3行（索引2）的标题
if len(data) &gt; 2:
    print(f"\n第3行标题（索引2）: {data[2]}")

# 保存为JSON以便后续分析
with open('excel_content.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print("\nExcel内容已保存到 excel_content.json")
