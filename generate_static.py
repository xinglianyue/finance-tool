"""
财务账单分析 - 纯静态版本
使用Python处理Excel，生成纯静态HTML，不依赖外部JS库
"""
import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime

# ============ 配置 ============
EXCEL_FILE = "3月份账单明细.xlsx"  # 或 "3月账单-新.xlsx"

# 城市分级
CITY_TIERS = {
    'E': ['承德市'],
    'F': ['围场满族蒙古族自治县', '玉田县', '安国市', '安平', '献县'],
    'G': ['晋州', '威县', '深泽县', '康保县']
}

# Excel行列映射 (1-based)
# 根据实际Excel结构
EXCEL_ROWS = {
    'CITY_NAME': 3,        # 城市名所在行
    'CITY_START_COL': 4,   # 城市列起始（承德市）
    'CITY_END_COL': 13,    # 城市列结束（康保县）
    'METRICS': {
        'orders': 7,       # 订单量
        'revenue': 37,     # 收入汇总
        'cost': 72,        # 成本汇总（负数）
        'profit': 73,      # 盈利（毛利）
    }
}

def get_tier(city_name):
    """获取城市分级"""
    for tier, cities in CITY_TIERS.items():
        for city in cities:
            if city in city_name:
                return tier
    return 'F'  # 默认F级

def parse_value(val):
    """解析单元格值"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').replace('¥', '').replace('元', ''))
    except:
        return 0

def process_excel():
    """处理Excel文件"""
    print(f"正在读取: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    cities = []
    totals = {'revenue': 0, 'cost': 0, 'profit': 0, 'orders': 0}

    # 遍历城市列
    for col in range(EXCEL_ROWS['CITY_START_COL'], EXCEL_ROWS['CITY_END_COL'] + 1):
        city_name = ws.cell(row=EXCEL_ROWS['CITY_NAME'], column=col).value
        if not city_name or str(city_name).strip() == '':
            continue

        orders = parse_value(ws.cell(row=EXCEL_ROWS['METRICS']['orders'], column=col).value)
        revenue = parse_value(ws.cell(row=EXCEL_ROWS['METRICS']['revenue'], column=col).value)
        cost = parse_value(ws.cell(row=EXCEL_ROWS['METRICS']['cost'], column=col).value)
        profit = parse_value(ws.cell(row=EXCEL_ROWS['METRICS']['profit'], column=col).value)

        # 计算UE和毛利率
        unit_ue = profit / orders if orders > 0 else 0
        margin = (revenue - cost) / revenue * 100 if revenue > 0 else 0

        city_data = {
            'name': str(city_name).strip(),
            'tier': get_tier(str(city_name)),
            'orders': orders,
            'revenue': revenue,
            'cost': cost,
            'profit': profit,
            'unit_ue': unit_ue,
            'margin': margin
        }
        cities.append(city_data)

        # 累加总计
        totals['revenue'] += revenue
        totals['cost'] += cost
        totals['profit'] += profit
        totals['orders'] += orders

    # 计算平均UE
    totals['avg_ue'] = totals['profit'] / totals['orders'] if totals['orders'] > 0 else 0
    totals['avg_margin'] = (totals['revenue'] - totals['cost']) / totals['revenue'] * 100 if totals['revenue'] > 0 else 0

    print(f"解析完成，共 {len(cities)} 个城市")
    return cities, totals

def generate_html(cities, totals):
    """生成静态HTML"""
    # 排序：亏损城市在前面
    sorted_cities = sorted(cities, key=lambda x: x['profit'])

    # 分类城市
    red_cities = [c for c in cities if c['profit'] < 0]
    orange_cities = [c for c in cities if 0 <= c['profit'] < 5000]
    green_cities = [c for c in cities if c['profit'] >= 5000]

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>财务账单分析 - {datetime.now().strftime('%Y年%m月')}</title>
<style>
:root {{
  --primary: #667eea;
  --success: #10b981;
  --warning: #f59e0b;
  --danger: #ef4444;
  --bg: #f0f2f5;
  --card: #ffffff;
  --text: #1e293b;
  --text-secondary: #64748b;
  --border: #e2e8f0;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: var(--bg); color: var(--text); padding: 16px; padding-bottom: 80px; }}
.header {{ background: var(--card); padding: 14px 20px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.08); border-radius: 12px; margin-bottom: 16px; }}
.header h1 {{ font-size: 18px; color: var(--primary); font-weight: 700; }}
.header .subtitle {{ font-size: 12px; color: var(--text-secondary); margin-top: 4px; }}

.cards {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 12px; margin-bottom: 16px; }}
.card {{ background: var(--card); border-radius: 12px; padding: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.card-label {{ font-size: 12px; color: var(--text-secondary); margin-bottom: 4px; }}
.card-value {{ font-size: 20px; font-weight: 700; }}
.card-value.success {{ color: var(--success); }}
.card-value.warning {{ color: var(--warning); }}
.card-value.danger {{ color: var(--danger); }}
.card-unit {{ font-size: 11px; color: var(--text-secondary); margin-top: 2px; }}

.alert-cards {{ display: flex; gap: 12px; margin-bottom: 16px; flex-wrap: wrap; }}
.alert-card {{ flex: 1; min-width: 100px; background: var(--card); border-radius: 12px; padding: 14px; text-align: center; cursor: pointer; box-shadow: 0 2px 8px rgba(0,0,0,0.06); transition: transform 0.2s; }}
.alert-card:active {{ transform: scale(0.98); }}
.alert-card.red {{ border-left: 3px solid var(--danger); }}
.alert-card.orange {{ border-left: 3px solid var(--warning); }}
.alert-card.green {{ border-left: 3px solid var(--success); }}
.alert-num {{ font-size: 24px; font-weight: 700; }}
.alert-card.red .alert-num {{ color: var(--danger); }}
.alert-card.orange .alert-num {{ color: var(--warning); }}
.alert-card.green .alert-num {{ color: var(--success); }}
.alert-label {{ font-size: 12px; color: var(--text-secondary); margin-top: 4px; }}

.section-title {{ font-size: 14px; font-weight: 600; color: var(--text-secondary); margin: 16px 0 10px; }}
.city-table {{ background: var(--card); border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.city-table table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
.city-table th {{ background: #f8f9fc; color: var(--text-secondary); font-weight: 600; font-size: 11px; text-transform: uppercase; padding: 10px 14px; text-align: left; }}
.city-table td {{ padding: 11px 14px; border-top: 1px solid var(--border); }}
.city-table tr:first-child td {{ border-top: none; }}
.name-cell {{ font-weight: 600; }}
.tier-badge {{ display: inline-block; padding: 1px 7px; border-radius: 10px; font-size: 11px; font-weight: 600; margin-right: 6px; }}
.tier-e {{ background: rgba(102,126,234,0.15); color: #597ef7; }}
.tier-f {{ background: rgba(245,158,11,0.15); color: #d97706; }}
.tier-g {{ background: rgba(16,185,129,0.15); color: #059669; }}
.ue-ok {{ color: var(--success); }}
.ue-warn {{ color: var(--warning); }}
.ue-bad {{ color: var(--danger); }}

.footer {{ text-align: center; padding: 20px; color: var(--text-secondary); font-size: 12px; }}
</style>
</head>
<body>

<div class="header">
  <h1>📊 财务账单分析</h1>
  <div class="subtitle">数据更新: {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
</div>

<div class="cards">
  <div class="card">
    <div class="card-label">总收入</div>
    <div class="card-value">{totals['revenue']:,.0f}</div>
    <div class="card-unit">元</div>
  </div>
  <div class="card">
    <div class="card-label">总成本</div>
    <div class="card-value" style="color: var(--text)">{totals['cost']:,.0f}</div>
    <div class="card-unit">元</div>
  </div>
  <div class="card">
    <div class="card-label">总毛利</div>
    <div class="card-value {'success' if totals['profit'] >= 0 else 'danger'}">{totals['profit']:,.0f}</div>
    <div class="card-unit">元</div>
  </div>
  <div class="card">
    <div class="card-label">订单量</div>
    <div class="card-value">{totals['orders']:,.0f}</div>
    <div class="card-unit">单</div>
  </div>
</div>

<div class="alert-cards">
  <div class="alert-card red">
    <div class="alert-num">{len(red_cities)}</div>
    <div class="alert-label">亏损城市</div>
  </div>
  <div class="alert-card orange">
    <div class="alert-num">{len(orange_cities)}</div>
    <div class="alert-label">低毛利城市</div>
  </div>
  <div class="alert-card green">
    <div class="alert-num">{len(green_cities)}</div>
    <div class="alert-label">盈利城市</div>
  </div>
</div>

<div class="section-title">城市明细（按毛利排序）</div>
<div class="city-table">
  <table>
    <thead>
      <tr>
        <th>城市</th>
        <th>订单量</th>
        <th>收入</th>
        <th>毛利</th>
        <th>UE</th>
      </tr>
    </thead>
    <tbody>
'''

    for c in sorted_cities:
        ue_class = 'ue-ok' if c['unit_ue'] >= 1 else ('ue-warn' if c['unit_ue'] >= 0 else 'ue-bad')
        ue_sign = '+' if c['unit_ue'] >= 0 else ''
        profit_class = 'success' if c['profit'] >= 0 else 'danger'
        profit_sign = '+' if c['profit'] >= 0 else ''

        html += f'''      <tr>
        <td class="name-cell">
          <span class="tier-badge tier-{c['tier'].lower()}">{c['tier']}类</span>
          {c['name']}
        </td>
        <td>{c['orders']:,.0f}</td>
        <td>{c['revenue']:,.0f}</td>
        <td style="color: {'var(--success)' if c['profit'] >= 0 else 'var(--danger)'}">{profit_sign}{c['profit']:,.0f}</td>
        <td class="{ue_class}">{ue_sign}{c['unit_ue']:.3f}</td>
      </tr>
'''

    html += '''    </tbody>
  </table>
</div>

<div class="footer">
  财务账单分析工具 v7 (纯静态版)
</div>

</body>
</html>'''

    return html

def main():
    cities, totals = process_excel()
    html = generate_html(cities, totals)

    output_file = "index_static.html"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"\n[OK] Generate OK: {output_file}")

    # 输出摘要
    print(f"\n=== Data Summary ===")
    print(f"Revenue: {totals['revenue']:,.0f}")
    print(f"Cost: {totals['cost']:,.0f}")
    print(f"Profit: {totals['profit']:,.0f}")
    print(f"Orders: {totals['orders']:,.0f}")
    print(f"Avg UE: {totals['avg_ue']:.3f}")
    print(f"Cities: {len(cities)}")

if __name__ == "__main__":
    main()
