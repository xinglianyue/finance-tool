#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出数据为原始Excel格式
从GitHub读取数据，导出为与原始外卖账单一致的Excel格式

使用方式：
  python export_excel.py                    # 从GitHub读取最新数据
  python export_excel.py 20260519           # 导出指定日期数据
  python export_excel.py 20260519 模板.xlsx  # 使用指定模板
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
import json
import requests
import base64
import sys
import os
import shutil
import warnings
from datetime import datetime
from copy import copy

warnings.filterwarnings('ignore')

GITHUB_REPO = 'xinglianyue/finance-tool'
GITHUB_API = f"https://api.github.com/repos/{GITHUB_REPO}/contents/shared-data.json"

def load_token():
    import json
    import os
    token_file = os.path.join(os.path.dirname(__file__), 'config-db.json')
    if os.path.exists(token_file):
        with open(token_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            return config.get('github', {}).get('token', '')
    return ''

CITY_ORDER = [
    "承德市", "围场满族蒙古族自治县", "玉田县", "安国市", "安平",
    "献县", "晋州", "威县", "深泽县", "康保县"
]

CITY_ID_MAP = {
    "承德市": 1103,
    "围场满族蒙古族自治县": 1289,
    "玉田县": 14818,
    "安国市": 1027,
    "安平": 580,
    "献县": 4448,
    "晋州": 573,
    "威县": 4091,
    "深泽县": 14122,
    "康保县": 14121,
}

def fetch_from_github():
    token = load_token()
    if not token:
        print("错误: 无法获取GitHub Token")
        return []
    
    headers = {'Authorization': f'token {token}'}
    try:
        resp = requests.get(GITHUB_API, headers=headers, timeout=30)
    except Exception:
        resp = requests.get(GITHUB_API, headers=headers, timeout=30, verify=False)
    
    if resp.status_code == 200:
        data = resp.json()
        decoded = json.loads(base64.b64decode(data['content']).decode('utf-8'))
        return decoded if isinstance(decoded, list) else [decoded]
    return []

def get_field_value(city_modules, field_name):
    if not city_modules:
        return None
    
    all_data = city_modules.get('all', {})
    return all_data.get(field_name)

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

def create_export_excel(data, output_path, template_path=None):
    if not data:
        print("错误: 没有数据可供导出")
        return False
    
    if template_path and os.path.exists(template_path):
        print(f"使用模板: {template_path}")
        shutil.copy2(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        template = r"C:\Users\xinxi\Desktop\美团工作\外卖账单20260519.xlsx"
        if os.path.exists(template):
            print(f"使用默认模板: {template}")
            shutil.copy2(template, output_path)
            wb = openpyxl.load_workbook(output_path)
        else:
            print("警告: 未找到模板文件，将创建空白Excel")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for record in data:
            merchant_data = record.get('merchantData', {})
            
            for block_key, block_info in [('all', '全量商家'), ('city', '城市商家'), ('ka', 'KA商家')]:
                if block_info != sheet_name:
                    continue
                
                block_data = merchant_data.get(block_key, {})
                cities = block_data.get('cities', [])
                
                if not cities:
                    continue
                
                print(f"处理 {sheet_name}: {len(cities)} 个城市")
                
                for city in cities:
                    city_name = city.get('name', '')
                    if city_name not in CITY_ORDER:
                        continue
                    
                    city_idx = CITY_ORDER.index(city_name)
                    data_col = 6 + city_idx
                    
                    modules = city.get('modules', {})
                    
                    for row_idx in range(5, ws.max_row + 1):
                        indicator = ws.cell(row_idx, 5).value
                        if not indicator:
                            continue
                        
                        field_map = {
                            '加盟原价交易额': 'franchiseGMV',
                            '自配原价交易额': 'selfGMV',
                            '原价交易额汇总': 'gmvAmount',
                            '加盟订单量': 'franchiseOrders',
                            '自配订单量': 'selfOrders',
                            '企客订单量': 'enterpriseOrders',
                            '订单量汇总': 'orders',
                            '加盟抽佣金额': 'franchiseCommission',
                            '自配抽佣金额': 'selfCommission',
                            '企客商家抽佣金额': 'enterpriseCommission',
                            '抽佣金额汇总': 'commission',
                            '加盟配送费': 'franchiseDeliveryFee',
                            '二次配送费': 'secondDeliveryFee',
                            '企客配送费': 'enterpriseDeliveryFee',
                            '一对一急送配送费': 'urgentDeliveryFee',
                            '配送费汇总': 'deliveryFee',
                            '合作商运营服务费': 'otherRevenue',
                            '拼单宝激励': 'pinDanReward',
                            '神券包激励': 'shenQuanReward',
                            '广告收入': 'adRevenue',
                            '专项补贴': 'specialSubsidy',
                            '星火激励调账': 'xingHuoAdjust',
                            '跑腿结算调账': 'runErrandAdjust',
                            '众包补贴调账': 'crowdSubsidyAdjust',
                            '竞价返还调账': 'biddingReturnAdjust',
                            '发展计划调账': 'developPlanAdjust',
                            '天补调账': 'weatherSubsidyAdjust',
                            '其他收入汇总': 'otherRevenueTotal',
                            '线上收入汇总': 'onlineRevenue',
                            '收入汇总': 'totalRevenue',
                            'B端代补金额': 'subsidyB',
                            'C端代补金额': 'subsidyC',
                            '账单-代补差额': 'subsidyDiff',
                            '代补金额花费汇总': 'subsidyTotal',
                            '平台抽佣金额': 'platformCommissionCost',
                            '合作商售后赔付费用': 'afterSaleCost',
                            '关爱基金': 'careFund',
                            '保险费用': 'insuranceCost',
                            '竞价': 'biddingCost',
                            '罚款': 'penalty',
                            'AI外呼费用结算': 'aiCallCost',
                            '平台成本汇总': 'platformCost',
                            '线上支出汇总': 'onlineExpense',
                            '办公室房租': 'officeRent',
                            '业务团队': 'teamCost',
                            '固定成本汇总': 'fixedCost',
                            '加盟承接订单量': 'franchiseDeliverOrders',
                            '加盟单均邮资': 'franchiseAvgPostage',
                            '加盟活动花费': 'franchiseActivityCost',
                            '加盟邮资': 'franchiseDelivery',
                            '普众众包订单量': 'crowdOrders',
                            '普众众包基础邮资': 'crowdBasePostage',
                            '普众众包活动花费': 'crowdActivityCost',
                            '普众众包邮资': 'crowdDelivery',
                            '悦跑订单量': 'yuepaoOrders',
                            '悦跑基础邮资': 'yuepaoBasePostage',
                            '悦跑周激励': 'yuepaoWeekReward',
                            '悦跑活动花费': 'yuepaoActivityCost',
                            '悦跑邮资': 'yuepaoDelivery',
                            '众包总邮资': 'crowdTotalDelivery',
                            '众包天气补贴': 'weatherSubsidy',
                            '配送成本汇总': 'deliveryCost',
                            '三方服务费': 'thirdPartyServiceCost',
                            '社保': 'socialInsurance',
                            '税': 'taxCost',
                            '附加成本汇总': 'additionalCost',
                            '外卖运营增单': 'operationBoostCost',
                            '水电电话网物料费': 'utilityCost',
                            '差旅招待': 'travelCost',
                            '其他成本': 'otherMiscCost',
                            '其他成本汇总': 'otherCost',
                            '线下支出汇总': 'offlineExpense',
                            '支出汇总': 'totalExpense',
                            '毛利': 'profit',
                        }
                        
                        field_name = field_map.get(indicator)
                        if field_name:
                            value = get_field_value(modules, field_name)
                            if value is not None and value != 0:
                                ws.cell(row_idx, data_col).value = value
    
    wb.save(output_path)
    print(f"导出成功: {output_path}")
    return True

def main():
    date_str = datetime.now().strftime('%Y%m%d')
    template_path = None
    
    if len(sys.argv) > 1:
        date_str = sys.argv[1]
    if len(sys.argv) > 2:
        template_path = sys.argv[2]
    
    print(f"正在从GitHub获取数据...")
    data = fetch_from_github()
    
    if not data:
        print("未获取到数据，请先使用 db-sync.py 同步数据")
        print("或者直接复制原始Excel文件:")
        print("  copy_excel.py")
        sys.exit(1)
    
    output_path = f"外卖账单导出{date_str}.xlsx"
    success = create_export_excel(data, output_path, template_path)
    
    if success:
        print(f"\n完成！导出文件: {output_path}")

if __name__ == "__main__":
    main()
