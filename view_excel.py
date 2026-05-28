
import openpyxl

file_path = '外卖账单20260527.xlsx'
wb = openpyxl.load_workbook(file_path)

print("=== 工作表列表 ===")
for sheet in wb.worksheets:
    print("-", sheet.title)
    print()

ws = wb.worksheets[0]
print(f"=== 工作表 '{ws.title}' 的前 20 行 ===")

for i in range(1, 21):
    row_data = []
    for j in range(1, 21):
        cell = ws.cell(row=i, column=j)
        value = cell.value
        if value is not None:
            row_data.append(str(value)[:20])
        else:
            row_data.append("")
    print(f"第 {i:2d} 行: {row_data}")

print(f"\n总行数: {ws.max_row}, 总列数: {ws.max_column}")
